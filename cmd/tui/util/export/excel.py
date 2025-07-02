import json
import os
import sys
import tempfile
from collections import Counter
from datetime import datetime
from datetime import datetime as dt

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_excel_file(tables):
    print("Starting create_excel_file", file=sys.stderr)
    # Initialize workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    print("Workbook initialized", file=sys.stderr)

    # Define styles
    header_fill = PatternFill(
        start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
    )
    header_font = Font(bold=True)
    even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for table in tables:
        table_name = table["name"]
        rows = table["rows"]
        print(f"Processing table: {table_name}, rows: {len(rows)}", file=sys.stderr)

        # Create data sheet
        data_sheet = wb.create_sheet(f"Data - {table_name}")
        print(f"Created data sheet: Data - {table_name}", file=sys.stderr)

        # Define headers
        if table_name == "Logs":
            headers = ["Time", "Level", "Message"]
        elif table_name == "Resource records":
            headers = ["ID", "Domain", "Data", "Type", "Class", "TTL"]
        elif table_name == "Users":
            headers = ["ID", "Login", "FirstName", "LastName", "Role"]
        else:
            headers = []
        print(f"Headers: {headers}", file=sys.stderr)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = data_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Write rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = data_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.fill = even_fill if row_idx % 2 == 0 else odd_fill
        print(f"Wrote {len(rows)} rows to data sheet", file=sys.stderr)

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row in data_sheet.iter_rows(
                min_row=2, max_row=len(rows) + 1, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            data_sheet.column_dimensions[col_letter].width = max_length + 2
        print("Adjusted column widths", file=sys.stderr)

        # Create chart sheet
        chart_sheet = wb.create_sheet(f"Charts - {table_name}")
        print(f"Created chart sheet: Charts - {table_name}", file=sys.stderr)

        if table_name == "Logs" and rows:
            # Pie Chart: Count by Level
            level_counts = Counter(row[1].upper() for row in rows if len(row) > 1)
            print(f"Level counts: {dict(level_counts)}", file=sys.stderr)
            if level_counts:
                chart_sheet["A1"] = "Level"
                chart_sheet["B1"] = "Count"
                levels = ["ERROR", "INFO", "WARN"]
                for idx, level in enumerate(levels, 2):
                    chart_sheet[f"A{idx}"] = level
                    chart_sheet[f"B{idx}"] = level_counts.get(level, 0)

                pie_chart = PieChart()
                pie_chart.title = "Messages by Level"
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=4)
                data = Reference(chart_sheet, min_col=2, min_row=1, max_row=4)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                chart_sheet.add_chart(pie_chart, "A5")
                print("Added pie chart for Levels", file=sys.stderr)

            # Column Chart: Messages by Day
            day_counts = Counter()
            for row in rows:
                if len(row) > 0:
                    try:
                        day = dt.strptime(row[0], "%Y-%m-%d %H:%M:%S").strftime(
                            "%Y-%m-%d"
                        )
                        day_counts[day] += 1
                    except ValueError as e:
                        print(
                            f"Skipping invalid date format in row {row}: {e}",
                            file=sys.stderr,
                        )
                        continue
            print(f"Day counts: {dict(day_counts)}", file=sys.stderr)
            if day_counts:
                days = sorted(day_counts.keys())
                chart_sheet["D1"] = "Day"
                chart_sheet["E1"] = "Count"
                for idx, day in enumerate(days, 2):
                    chart_sheet[f"D{idx}"] = day
                    chart_sheet[f"E{idx}"] = day_counts[day]

                bar_chart = BarChart()
                bar_chart.title = "Messages by Day"
                bar_chart.type = "col"
                bar_chart.dataLabels = DataLabelList()
                bar_chart.dataLabels.showVal = True
                labels = Reference(
                    chart_sheet, min_col=4, min_row=2, max_row=len(days) + 1
                )
                data = Reference(
                    chart_sheet, min_col=5, min_row=1, max_row=len(days) + 1
                )
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(labels)
                chart_sheet.add_chart(bar_chart, "I5")
                print("Added column chart for Days", file=sys.stderr)

        elif table_name == "Resource records" and rows:
            # Pie Chart: Count by Type
            type_counts = Counter(row[3].upper() for row in rows if len(row) > 3)
            print(f"Type counts: {dict(type_counts)}", file=sys.stderr)
            if type_counts:
                types = sorted(type_counts.keys())
                chart_sheet["A1"] = "Type"
                chart_sheet["B1"] = "Count"
                for idx, typ in enumerate(types, 2):
                    chart_sheet[f"A{idx}"] = typ
                    chart_sheet[f"B{idx}"] = type_counts[typ]

                pie_chart = PieChart()
                pie_chart.title = "Records by Type"
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                labels = Reference(
                    chart_sheet, min_col=1, min_row=2, max_row=len(types) + 1
                )
                data = Reference(
                    chart_sheet, min_col=2, min_row=1, max_row=len(types) + 1
                )
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                chart_sheet.add_chart(pie_chart, "A5")
                print("Added pie chart for Types", file=sys.stderr)

            # Pie Chart: Count by Class
            class_counts = Counter(row[4].upper() for row in rows if len(row) > 4)
            print(f"Class counts: {dict(class_counts)}", file=sys.stderr)
            if class_counts:
                classes = sorted(class_counts.keys())
                chart_sheet["D1"] = "Class"
                chart_sheet["E1"] = "Count"
                for idx, cls in enumerate(classes, 2):
                    chart_sheet[f"D{idx}"] = cls
                    chart_sheet[f"E{idx}"] = class_counts[cls]

                pie_chart = PieChart()
                pie_chart.title = "Records by Class"
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                print("Adding pie chart for Classes", file=sys.stderr)
                labels = Reference(
                    chart_sheet, min_col=4, min_row=2, max_row=len(classes) + 1
                )
                data = Reference(
                    chart_sheet, min_col=5, min_row=1, max_row=len(classes) + 1
                )
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                chart_sheet.add_chart(pie_chart, "I5")
                print("Added pie chart for Classes", file=sys.stderr)

        elif table_name == "Users" and rows:
            # Pie Chart: Count by Role
            role_counts = Counter(row[4].upper() for row in rows if len(row) > 8)
            print(f"Role counts: {dict(role_counts)}", file=sys.stderr)
            if role_counts:
                roles = sorted(role_counts.keys())
                chart_sheet["A1"] = "Role"
                chart_sheet["B1"] = "Count"
                for idx, role in enumerate(roles, 2):
                    chart_sheet[f"A{idx}"] = role
                    chart_sheet[f"B{idx}"] = role_counts[role]

                pie_chart = PieChart()
                pie_chart.title = "Users by Role"
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                labels = Reference(
                    chart_sheet, min_col=1, min_row=2, max_row=len(roles) + 1
                )
                data = Reference(
                    chart_sheet, min_col=2, min_row=1, max_row=len(roles) + 1
                )
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                chart_sheet.add_chart(pie_chart, "A5")
                print("Added pie chart for Roles", file=sys.stderr)

    # Save file to temporary directory
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f"consolidated_report_{timestamp}.xlsx"
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, file_name)
    print(f"Saving file to: {file_path}", file=sys.stderr)
    wb.save(file_path)
    print(f"File saved successfully: {file_path}", file=sys.stderr)
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} was not created", file=sys.stderr)
        raise FileNotFoundError(f"File {file_path} was not created")
    return file_path


if __name__ == "__main__":
    try:
        print("Reading JSON input", file=sys.stderr)
        input_data = json.load(sys.stdin)
        print(f"Input data: {input_data}", file=sys.stderr)
        file_path = create_excel_file(input_data)
        print(file_path)  # Ensure file path is printed to stdout
    except Exception as e:
        print(f"Error in Python script: {str(e)}", file=sys.stderr)
        sys.exit(1)
